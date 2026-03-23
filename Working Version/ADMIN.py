from openpyxl import load_workbook
import hashlib

INVENTORY_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCons Cart Inventory.xlsx"
RENTALS_FILE   = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCon Equipment Rentals 2024-26.xlsx"

# Rentals column indices (0-based)
COL_ICON_NAME     = 0
COL_DATE          = 1
COL_FIRST_NAME    = 2
COL_LAST_NAME     = 3
COL_EQUIPMENT     = 4
COL_AMOUNT        = 5
COL_CARD_TAKEN    = 6
COL_CARD_RETURNED = 7
COL_SIGNED_OUT    = 8
COL_SIGNED_IN     = 9
COL_COMMENTS      = 10

RENTALS_HEADERS = [
    "iCon Name", "Date", "First Name", "Last Name", "Equipment",
    "Amount", "Student Card Taken", "Student Card Returned",
    "Signed Out", "Signed In", "Comments"
]


# ── Inventory ────────────────────────────────────────────────────────────────

def add_equipment(name: str, quantity: int):
    """
    Adds a new item to inventory, or increases quantity if it already exists.
    """
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            row[1].value = (row[1].value or 0) + quantity
            wb.save(INVENTORY_FILE)
            return
    ws.append([name, quantity])
    wb.save(INVENTORY_FILE)


def remove_equipment(name: str):
    """
    Removes an item row entirely from inventory.
    """
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == name:
            ws.delete_rows(i)
            wb.save(INVENTORY_FILE)
            return
    raise ValueError(f"'{name}' not found in inventory.")


def get_all_equipment():
    """
    Returns list of (name, quantity) tuples for the remove dropdown.
    """
    try:
        wb = load_workbook(INVENTORY_FILE)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                items.append((row[0], row[1] or 0))
        return items
    except Exception as e:
        print(f"Error reading inventory: {e}")
        return []


# ── iCon accounts ────────────────────────────────────────────────────────────

def _hash(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def add_icon(username: str, password: str):
    with open("iData.txt", "r") as f:
        for line in f:
            if line.strip().split(",")[0] == username:
                raise ValueError(f"Username '{username}' already exists.")
    with open("iData.txt", "a") as f:
        f.write(f"{username},{_hash(password)}\n")


def remove_icon(username: str):
    with open("iData.txt", "r") as f:
        lines = f.readlines()
    new_lines = [l for l in lines if l.strip().split(",")[0] != username]
    if len(new_lines) == len(lines):
        raise ValueError(f"Username '{username}' not found.")
    with open("iData.txt", "w") as f:
        f.writelines(new_lines)


def get_icon_usernames():
    try:
        with open("iData.txt", "r") as f:
            return [line.strip().split(",")[0] for line in f if line.strip()]
    except Exception:
        return []


# ── Records ──────────────────────────────────────────────────────────────────

def get_all_records():
    """
    Returns a list of dicts for every non-empty row in the rentals file.
    Each dict also includes 'row_number' (1-based Excel row).
    """
    try:
        wb = load_workbook(RENTALS_FILE, data_only=True)
        ws = wb.active
        records = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[COL_FIRST_NAME] and not row[COL_EQUIPMENT]:
                continue
            records.append({
                "row_number":      i,
                "icon_name":       row[COL_ICON_NAME]     or "",
                "date":            row[COL_DATE]           or "",
                "first_name":      row[COL_FIRST_NAME]     or "",
                "last_name":       row[COL_LAST_NAME]      or "",
                "equipment":       row[COL_EQUIPMENT]      or "",
                "amount":          row[COL_AMOUNT]         or 0,
                "card_taken":      row[COL_CARD_TAKEN],
                "card_returned":   row[COL_CARD_RETURNED],
                "signed_out":      row[COL_SIGNED_OUT],
                "signed_in":       row[COL_SIGNED_IN],
                "comments":        row[COL_COMMENTS]       or "",
            })
        return records
    except Exception as e:
        print(f"Error reading records: {e}")
        return []


def update_record(row_number: int, field: str, new_value):
    """
    Updates a single field in a specific row of the rentals file.
    field must be one of the RENTALS_HEADERS values.
    """
    col_map = {
        "iCon Name":             COL_ICON_NAME,
        "Date":                  COL_DATE,
        "First Name":            COL_FIRST_NAME,
        "Last Name":             COL_LAST_NAME,
        "Equipment":             COL_EQUIPMENT,
        "Amount":                COL_AMOUNT,
        "Student Card Taken":    COL_CARD_TAKEN,
        "Student Card Returned": COL_CARD_RETURNED,
        "Signed Out":            COL_SIGNED_OUT,
        "Signed In":             COL_SIGNED_IN,
        "Comments":              COL_COMMENTS,
    }
    if field not in col_map:
        raise ValueError(f"Unknown field '{field}'")

    wb = load_workbook(RENTALS_FILE)
    ws = wb.active
    col_index = col_map[field] + 1  # openpyxl is 1-based
    ws.cell(row=row_number, column=col_index).value = new_value
    wb.save(RENTALS_FILE)


def delete_record(row_number: int):
    """
    Deletes a row from the rentals file entirely.
    """
    wb = load_workbook(RENTALS_FILE)
    ws = wb.active
    ws.delete_rows(row_number)
    wb.save(RENTALS_FILE)