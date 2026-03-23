from datetime import datetime
from openpyxl import load_workbook

RENTALS_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCon Equipment Rentals 2024-26.xlsx"
INVENTORY_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCons Cart Inventory.xlsx"

# Column indices in the rentals sheet (0-based)
COL_ICON_NAME       = 0
COL_DATE            = 1
COL_FIRST_NAME      = 2
COL_LAST_NAME       = 3
COL_EQUIPMENT       = 4
COL_AMOUNT          = 5
COL_CARD_TAKEN      = 6
COL_CARD_RETURNED   = 7
COL_SIGNED_OUT      = 8
COL_SIGNED_IN       = 9
COL_COMMENTS        = 10


def get_open_loans():
    """
    Returns a list of display strings for all rows where Signed In is False.
    Format: "Firstname Lastname — Equipment (x Amount)  [row N]"
    The row number is embedded so process_return can find the exact row.
    """
    try:
        wb = load_workbook(RENTALS_FILE, data_only=True)
        ws = wb.active
        loans = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            first = row[COL_FIRST_NAME]
            equip = row[COL_EQUIPMENT]
            # Skip completely empty rows
            if not first and not equip:
                continue
            signed_in = row[COL_SIGNED_IN]
            if signed_in is False or signed_in == 0:
                last   = row[COL_LAST_NAME] or ""
                amount = row[COL_AMOUNT]    or ""
                label  = f"{first} {last} — {equip} (x{amount})  [row {i}]"
                loans.append(label)
        return loans
    except Exception as e:
        print(f"Error reading rentals: {e}")
        return []


def process_return(loan_label: str, card_returned: bool, comments: str):
    """
    Called by the GUI when a return is submitted.
    - Marks Signed In = True and Card Returned on the matching row
    - Adds the return timestamp to Comments
    - Adds the returned quantity back to inventory
    """
    row_number = _parse_row_number(loan_label)
    if row_number is None:
        raise ValueError("Could not identify the loan record from the selection.")

    wb = load_workbook(RENTALS_FILE)
    ws = wb.active

    row = ws[row_number]
    equipment = row[COL_EQUIPMENT].value
    amount    = row[COL_AMOUNT].value or 0

    now = datetime.now().strftime("%d/%m/%Y %I:%M %p")
    existing_comments = row[COL_COMMENTS].value or ""
    new_comments = f"{existing_comments} | Returned: {now}" if existing_comments else f"Returned: {now}"
    if comments:
        new_comments += f" | Note: {comments}"

    row[COL_SIGNED_IN].value      = True
    row[COL_CARD_RETURNED].value  = card_returned
    row[COL_COMMENTS].value       = new_comments

    wb.save(RENTALS_FILE)

    _add_to_inventory(equipment, amount)


def _parse_row_number(label: str):
    """
    Extracts the row number embedded in the loan label string.
    e.g. "John Smith — Laptop (x2)  [01/01/2025 ...] [row 5]" → 5
    """
    try:
        start = label.rfind("[row ") + 5
        end   = label.rfind("]")
        return int(label[start:end].strip())
    except Exception:
        return None


def _add_to_inventory(equipment: str, amount: int):
    """
    Adds the returned quantity back to the inventory file.
    """
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == equipment:
            row[1].value = (row[1].value or 0) + amount
            break
    wb.save(INVENTORY_FILE)