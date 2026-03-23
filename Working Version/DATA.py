from openpyxl import load_workbook
import json
import os

RENTALS_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCon Equipment Rentals 2024-26.xlsx"
OUT_OF_STOCK_LOG = "out_of_stock_log.json"

# Rentals column indices (0-based)
COL_EQUIPMENT = 4
COL_AMOUNT    = 5
COL_SIGNED_IN = 9


def get_table_data():
    """
    Returns a list of dicts, one per item ever loaned:
    {
        "item":             str,
        "borrow_count":     int,  # number of loan transactions
        "total_loaned":     int,  # total quantity loaned across all transactions
        "currently_on_loan":int,  # quantity where Signed In is still False
        "out_of_stock":     int,  # times item hit 0 stock (from log)
    }
    Sorted alphabetically by item name.
    """
    borrow_count      = {}
    total_loaned      = {}
    currently_on_loan = {}

    try:
        wb = load_workbook(RENTALS_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            equip = row[COL_EQUIPMENT]
            if not equip:
                continue
            try:
                amount = int(row[COL_AMOUNT] or 0)
            except (ValueError, TypeError):
                amount = 0
            signed_in = row[COL_SIGNED_IN]

            borrow_count[equip]       = borrow_count.get(equip, 0) + 1
            total_loaned[equip]       = total_loaned.get(equip, 0) + amount
            if signed_in is False or signed_in == 0:
                currently_on_loan[equip] = currently_on_loan.get(equip, 0) + amount

    except Exception as e:
        print(f"Error reading rentals: {e}")

    oos_counts = _load_out_of_stock_log()

    all_items = sorted(set(list(borrow_count.keys()) + list(oos_counts.keys())))

    return [
        {
            "item":              item,
            "borrow_count":      borrow_count.get(item, 0),
            "total_loaned":      total_loaned.get(item, 0),
            "currently_on_loan": currently_on_loan.get(item, 0),
            "out_of_stock":      oos_counts.get(item, 0),
        }
        for item in all_items
    ]


def log_out_of_stock(item: str):
    """
    Called by LOAN.py whenever an item reaches 0 after a loan.
    Increments the count for that item in the log file.
    """
    log = _load_out_of_stock_log()
    log[item] = log.get(item, 0) + 1
    _save_out_of_stock_log(log)


def _load_out_of_stock_log():
    if os.path.exists(OUT_OF_STOCK_LOG):
        try:
            with open(OUT_OF_STOCK_LOG, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_out_of_stock_log(log: dict):
    with open(OUT_OF_STOCK_LOG, "w") as f:
        json.dump(log, f, indent=2)