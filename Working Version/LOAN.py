from datetime import datetime
from openpyxl import load_workbook
from DATA import log_out_of_stock

RENTALS_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCon Equipment Rentals 2024-26.xlsx"
INVENTORY_FILE = "C:/Users/markb/OneDrive/Documents/APSC 103/iCons Folders/iCons Cart Inventory.xlsx"


def get_equipment_list():
    """
    Returns only items with stock > 0.
    """
    in_stock, _ = get_equipment_list_with_stock()
    return in_stock


def get_equipment_list_with_stock():
    """
    Returns two lists: (in_stock, out_of_stock)
    in_stock     — item names where quantity > 0
    out_of_stock — item names where quantity <= 0
    """
    try:
        wb = load_workbook(INVENTORY_FILE)
        ws = wb.active
        in_stock = []
        out_of_stock = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                qty = row[1] or 0
                if qty > 0:
                    in_stock.append(row[0])
                else:
                    out_of_stock.append(row[0])
        return in_stock, out_of_stock
    except Exception as e:
        print(f"Error reading inventory: {e}")
        return [], []


def process_loan(data: dict):
    """
    Main function called by GUI.
    Writes the loan record to the rentals file and subtracts from inventory.
    Raises an exception (with a message) if there is not enough stock.
    """
    _check_inventory(data["equipment"], data["amount"])
    record = _build_record(data)
    _write_to_rentals(record)
    _subtract_inventory(data["equipment"], data["amount"])


def _check_inventory(equipment: str, amount: int):
    """
    Raises ValueError if there is not enough stock.
    """
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == equipment:
            current_qty = row[1] or 0
            if amount > current_qty:
                raise ValueError(
                    f"Not enough stock for '{equipment}'. "
                    f"Requested: {amount}, Available: {current_qty}"
                )
            return
    raise ValueError(f"Equipment '{equipment}' not found in inventory.")


def _subtract_inventory(equipment: str, amount: int):
    """
    Subtracts the loaned amount from the inventory file.
    Logs an out-of-stock event if the item hits 0 after the loan.
    """
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == equipment:
            new_qty = (row[1].value or 0) - amount
            row[1].value = new_qty
            if new_qty == 0:
                log_out_of_stock(equipment)
            break
    wb.save(INVENTORY_FILE)


def _build_record(data: dict):
    now = datetime.now()
    formatted_time = now.strftime("%d/%m/%Y %I:%M %p")
    return {
        "iCon Name": data["icon_name"],
        "Date": formatted_time,
        "First Name": data["first_name"],
        "Last Name": data["last_name"],
        "Equipment": data["equipment"],
        "Amount": data["amount"],
        "Student Card Taken": data["student_card_taken"],
        "Student Card Returned": False,
        "Signed Out": data["student_card_taken"],
        "Signed In": False,
        "Comments": ""
    }


def _write_to_rentals(record: dict):
    wb = load_workbook(RENTALS_FILE)
    ws = wb.active
    ws.append([
        record["iCon Name"],
        record["Date"],
        record["First Name"],
        record["Last Name"],
        record["Equipment"],
        record["Amount"],
        record["Student Card Taken"],
        record["Student Card Returned"],
        record["Signed Out"],
        record["Signed In"],
        record["Comments"]
    ])
    wb.save(RENTALS_FILE)